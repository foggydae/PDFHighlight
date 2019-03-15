# coding = utf-8

import re
import xlrd
import os
import json
import csv
from openpyxl import load_workbook
from pprint import pprint

kywdFile = xlrd.open_workbook('new_alias.xlsx')
kywdTable = kywdFile.sheets()[0]
kywdRows = kywdTable.nrows
kywdCols = kywdTable.ncols

aliasDict = {}

for rowIndex in range(1, kywdRows):
	rowValue = kywdTable.row_values(rowIndex)

	if rowValue[1] not in aliasDict:
		aliasDict[rowValue[1]] = [rowValue[0]]
	elif rowValue[0] not in aliasDict[rowValue[1]]:
		aliasDict[rowValue[1]].append(rowValue[0])
	else:
		pass

wb = load_workbook('dict_with_alias.xlsx')
ws = wb['Sheet1']
for i in range(2, len(list(ws.rows)) + 1):
	thisNewAlias = []
	for j in range(3, 0, -1):
		if not ws.cell(row = i, column = j).value == None:
			curCode = ws.cell(row = i, column = j).value
			if curCode in aliasDict:
				thisNewAlias += aliasDict[curCode]
	thisNewAlias = set(thisNewAlias)
	startColIndex = 5
	for j in range(len(list(ws.rows)[i - 1]), 4, -1):
		if not ws.cell(row = i, column = j).value == None:
			startColIndex = j + 1
			break
	for j in range(startColIndex, startColIndex + len(thisNewAlias)):
		ws.cell(row = i, column = j, value = thisNewAlias.pop())
wb.save('dict_with_alias.xlsx')
